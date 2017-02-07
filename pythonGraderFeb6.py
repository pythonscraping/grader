from openpyxl import load_workbook
from tkinter import * 
from collections import defaultdict
import pickle 
import glob

#Dictionary containing all our data
dicc = {}
#Dictionary containing our Tkinter objects
labeldicc = {}
#Dictionary containing our stringvar objects 
stringvardicc = {}

listofexcelfiles = glob.glob("*.xlsx")


for file in listofexcelfiles:
    wb2 = load_workbook(file, read_only = True, data_only=True)
    wb1 = load_workbook(file, read_only = True)
    listofsheetnames = wb2.get_sheet_names()
    print(len(listofsheetnames), " are the sheet names")

for index,name in enumerate(listofsheetnames):
    print (index, " ", name)

dicc["listofsheetnames"] = listofsheetnames

print("Next Step")


#We define the cells we want to grade for each sheet. 
def listofcells():
    fenetre = Tk()
    currentrow = 0
    for index,name in enumerate(dicc["listofsheetnames"]):
        stringvardicc[index,"listofcells"] = StringVar()
        stringvardicc[index,"listofcells"].set("")
        labeldicc[index,"listofcells","displaytext"] = Label(fenetre, text="This is the sheet number " +(str(index+1))).grid(row=currentrow, column=0)
        labeldicc[index,"listofcells"]     = Entry(fenetre, text=stringvardicc[index,"listofcells"]).grid(row=currentrow+1, column=0)
        currentrow = currentrow + 2

    def on_closing():
        saveCells()
        saveUnique()
        fenetre.destroy() 
        

    def saveCells():
        for index,name in enumerate(dicc["listofsheetnames"]):
            dicc[index,"listofcells"] = stringvardicc[index,"listofcells"].get()
            for file in listofexcelfiles:
                wb2 = load_workbook(file, read_only = True, data_only=True)
                wb1 = load_workbook(file, read_only = True)
                a = wb2.get_sheet_names()
                #We only deal with the first sheet here
                b =  wb2[ a[index] ]
                b1 =  wb1[ a[index] ]

                for cell in dicc[index,"listofcells"].split(','):
                    cell = cell.strip()
                    print(cell,index," ",file, "TRACKING")
                    try :
                        dicc[index,file,cell,"value"] = b1[cell].value
                        dicc[index,file,cell,"formula"] = b[cell].value
                    except :
                        dicc[index,file,cell,"value"] = ""
                        dicc[index,file,cell,"formula"] = ""
        pickle.dump( dicc, open( "dicc.p", "wb" ))


    def saveUnique():
        unique = {}
        unique["listofsheetnames"] = dicc["listofsheetnames"]
        for index,name in enumerate(dicc["listofsheetnames"]):
            for cell in dicc[index,"listofcells"].split(','):
                cell = cell.strip()
                unique[index,cell] = []
            unique[index,"listofcells"] = dicc[index,"listofcells"] 
        for index,name in enumerate(dicc["listofsheetnames"]):        
            for file in listofexcelfiles:
                for cell in dicc[index,"listofcells"].split(','):
                    cell = cell.strip()
                    if (dicc[index,file,cell,"value"],dicc[index,file,cell,"formula"]) not in unique[index,cell]:
                        unique[index,cell].append((dicc[index,file,cell,"value"],dicc[index,file,cell,"formula"]))
        pickle.dump( unique, open( "unique.p", "wb" ))

    #Window configuration
    fenetre.configure(background='white')    
    fenetre.protocol("WM_DELETE_WINDOW", on_closing)
    w, h = fenetre.winfo_screenwidth(), fenetre.winfo_screenheight()
    fenetre.geometry("%dx%d+0+0" % (w, h))
    fenetre.mainloop()

listofcells()