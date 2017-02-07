from openpyxl import load_workbook
from tkinter import * 
from collections import defaultdict
import pickle 
import glob


dicc = pickle.load( open( "dicc.p", "rb" ) )
unique = pickle.load( open( "unique.p", "rb" ) )
listofexcelfiles = glob.glob("*.xlsx")
listofsheetnames = unique["listofsheetnames"]

for file in listofexcelfiles:
	print("Grading file : ", file)
	for indexofSheet,sheet in enumerate(listofsheetnames):
		print("Grading sheet number : ",indexofSheet)
		for cell in unique[indexofSheet,"listofcells"].split(","):
			#print("Grading cell : ",cell)
			couple = dicc[indexofSheet,file,cell,"value"],dicc[indexofSheet,file,cell,"formula"]
			for index,element in enumerate(unique[indexofSheet,cell]):
				if couple == element :
					couplePoints = unique[indexofSheet,cell,'numberofpoints']
					coupleGrade = unique[indexofSheet,cell,index,"grade"]
			print(cell," : ", coupleGrade, " / ", couplePoints)
	print()