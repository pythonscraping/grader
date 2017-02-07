from openpyxl import load_workbook
from tkinter import * 
from collections import defaultdict
import pickle 
import glob



unique = pickle.load( open( "unique.p", "rb" ) )

labeldicc =  {}
stringvardicc = {}




def gradeCell(sheet):


    global currentcell
    fenetre = Tk()
    listofcells = unique[sheet,"listofcells"].split(",")

    

    cell = listofcells[currentcell].strip()
    currentrow = 0
    introText = "Grading cell " + cell + " from sheet number " + str(sheet+1)
    labeldicc[sheet,0,cell,"cell"] = Label(fenetre, text=introText).grid(row=currentrow,column=1,pady=20)
    currentrow = currentrow + 1
    stringvardicc[sheet,cell,"numberofpoints"] = StringVar()
    try:
    	stringvardicc[sheet,cell,"numberofpoints"].set(unique[sheet,cell,"numberofpoints"])
    except KeyError:
    	stringvardicc[sheet,cell,"numberofpoints"].set("")
    labeldicc["display",0] = Label(fenetre, text="Number of points :").grid(row=currentrow, padx=10,column=0)
    labeldicc[sheet,cell,"numberofpoints"] = Entry(fenetre, textvariable=stringvardicc[sheet,cell,"numberofpoints"], width=30).grid(row=currentrow,column=1)
    currentrow = currentrow + 2
    labeldicc["display",0] = Label(fenetre, text="formula").grid(row=currentrow, padx=10,column=0)
    labeldicc["display",1] = Label(fenetre, text="value").grid(row=currentrow, padx=10,column=1)
    labeldicc["display",2] = Label(fenetre, text="grade").grid(row=currentrow, padx=10,column=2)
    currentrow = currentrow + 1
    for elementindex, element in enumerate(unique[sheet,cell]):
            labeldicc[sheet,1,cell,elementindex,"formula"] = Label(fenetre, text=unique[sheet,cell][elementindex][0]).grid(row=currentrow, padx=10,column=0)
            labeldicc[sheet,2,cell,elementindex,"formula"] = Label(fenetre, text=unique[sheet,cell][elementindex][1]).grid(row=currentrow,padx=10,column=1)
            stringvardicc[sheet,cell,elementindex,"grade"] = StringVar()
            try:
            	stringvardicc[sheet,cell,elementindex,"grade"].set(unique[sheet,cell,elementindex,"grade"])
            except KeyError:
            	stringvardicc[sheet,cell,elementindex,"grade"].set("")
            labeldicc[sheet,3,cell,elementindex,"formula"] = Entry(fenetre, textvariable=stringvardicc[sheet,cell,elementindex,"grade"], width=30).grid(row=currentrow,column=2)
            
            

            currentrow = currentrow + 1
    def nextcell():
        on_closing() 
        global currentcell
        currentcell = currentcell + 1
        if currentcell < len(listofcells):
        	gradeCell(sheet) 

    b = Button(fenetre, text="nextcell", command=nextcell).grid(column=1,pady=25)  
      
    def on_closing():
        savegrades()
        fenetre.destroy()  

    def savegrades():
        for elementindex, element in enumerate(unique[sheet,cell]):
             unique[sheet,cell,elementindex,"grade"]=stringvardicc[sheet,cell,elementindex,"grade"].get()
        unique[sheet,cell,"numberofpoints"]=stringvardicc[sheet,cell,"numberofpoints"].get()
        pickle.dump( unique, open( "unique.p", "wb" ) )


    #Window configuration
    fenetre.configure(background='white')    
    fenetre.protocol("WM_DELETE_WINDOW", on_closing)
    w, h = fenetre.winfo_screenwidth(), fenetre.winfo_screenheight()
    fenetre.geometry("%dx%d+0+0" % (w, h))
    fenetre.mainloop()


listofsheetnames = unique["listofsheetnames"]
for index2,sheet in enumerate(listofsheetnames):
	currentcell = 0
	gradeCell(index2)