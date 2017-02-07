from openpyxl import load_workbook
from tkinter import * 
from collections import defaultdict
import pickle 
import glob

dicc = {}
listofcells = ["C8", "D8", "C9","D9","C21"]



listofexcelfiles = glob.glob("*.xlsx")

#We get all the values in a dictionary 
for file in listofexcelfiles:
	wb2 = load_workbook(file, read_only = True, data_only=True)
	wb1 = load_workbook(file, read_only = True)
	a = wb2.get_sheet_names()
	#We only deal with the first sheet here
	b =  wb2[ a[0] ]
	b1 =  wb1[ a[0] ]

	for index,cell in enumerate(listofcells) :
		dicc[file,cell,"value"] = b1[cell].value
		dicc[file,cell,"formula"] = b[cell].value

#Now we group identical values in a new dictionary
unique = {}

for cell in listofcells:
		unique[cell] = []

for file in listofexcelfiles:
	for cell in listofcells:
		if (dicc[file,cell,"value"],dicc[file,cell,"formula"]) not in unique[cell]:
			unique[cell].append((dicc[file,cell,"value"],dicc[file,cell,"formula"]))


#Then we assign a 0 grade by default to every couple
for cell in listofcells:
	for index,element in enumerate(unique[cell]):
		unique[cell,index] = ""

import pprint
pp = pprint.PrettyPrinter(indent=4)

pickle.dump( dicc, open( "dicc.p", "wb" ) )
pickle.dump( unique, open( "unique.p", "wb" ) )
pickle.dump( listofcells, open( "listofcells.p", "wb" ) )
